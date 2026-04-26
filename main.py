# -*- coding: utf-8 -*-
"""
Thesis-ready statistical analysis for weightlifting dataset 2025.

Hypotheses covered:
H1: Vertical jump height (CMJ, SJ) vs competitive performance.
H2: Dominant handgrip strength vs competitive success, including Snatch vs Clean & Jerk comparison.
H3: Arm and thigh circumferences vs competitive performance, absolute vs relative comparison.
H4: Elite athletes (top 25% within each sex by Sinclair) vs others in RLL and RAL.
H5: Eccentric Utilization Ratio (EUR = CMJ / SJ) vs Sinclair score.

Outputs:
  thesis_outputs/
    weightlifting_thesis_results.xlsx
    publication_tables/*.csv
    figures/H1/*.png ... figures/H5/*.png
    thesis_interpretation_text.txt
    analysis_log.txt

Install dependencies if needed:
  pip install pandas numpy scipy openpyxl matplotlib

How to run:
  1) Put this script in the same folder as weightlifting_dataset_2025.csv
  2) Run: python weightlifting_thesis_ready_analysis.py
"""

from pathlib import Path
import re
import textwrap
import warnings

import numpy as np
import pandas as pd
from scipy import stats
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# =========================
# 1) Settings
# =========================
INPUT_FILE = Path("weightlifting_dataset_2025.csv")
OUTPUT_DIR = Path("thesis_outputs")
FIGURE_DIR = OUTPUT_DIR / "figures"
TABLE_DIR = OUTPUT_DIR / "publication_tables"
OUTPUT_XLSX = OUTPUT_DIR / "weightlifting_thesis_results.xlsx"
INTERPRETATION_TXT = OUTPUT_DIR / "thesis_interpretation_text.txt"
LOG_TXT = OUTPUT_DIR / "analysis_log.txt"

ALPHA = 0.05
RANDOM_SEED = 42
np.random.seed(RANDOM_SEED)

# Sinclair 2025-2028 parameters from the assignment
SINCLAIR_PARAMS = {
    "Muži": {"A": 0.722762521, "b": 193.609, "label": "Men"},
    "Ženy": {"A": 0.787004341, "b": 153.757, "label": "Women"},
}

EN_LABELS = {
    "Muži": "Men",
    "Ženy": "Women",
    "Junior": "U20",
    "U23": "U23",
    "B.weight": "Body mass (kg)",
    "Snatch": "Snatch (kg)",
    "Clean Jerk": "Clean & Jerk (kg)",
    "Result": "Total (kg)",
    "Sinclair": "Sinclair score",
    "Snatch Sinclair": "Snatch Sinclair score",
    "Clean Jerk Sinclair": "Clean & Jerk Sinclair score",
    "Total Sinclair calculated": "Total Sinclair score (calculated)",
    "CMJ Height": "CMJ height (cm)",
    "SJ Height": "SJ height (cm)",
    "Dominant handgrip": "Dominant handgrip strength (kg)",
    "Biceps (cm)": "Arm circumference (cm)",
    "Stehno (cm)": "Thigh circumference (cm)",
    "Výška (cm)": "Body height (cm)",
    "Délka nohy (cm)": "Leg length (cm)",
    "Délka ruky (cm)": "Arm length (cm)",
    "EUR": "Eccentric utilization ratio (EUR)",
    "RLL": "Relative leg length (RLL, %)",
    "RAL": "Relative arm length (RAL, %)",
    "Elite": "Elite group",
    "Others": "Others",
}

OUTPUT_DIR.mkdir(exist_ok=True)
FIGURE_DIR.mkdir(exist_ok=True)
TABLE_DIR.mkdir(exist_ok=True)
for h in ["H1", "H2", "H3", "H4", "H5"]:
    (FIGURE_DIR / h).mkdir(parents=True, exist_ok=True)

# =========================
# 2) Helper functions
# =========================
def to_numeric_series(s: pd.Series) -> pd.Series:
    """Convert strings with decimal comma or units to numeric."""
    if pd.api.types.is_numeric_dtype(s):
        return s
    return (
        s.astype(str)
        .str.replace(",", ".", regex=False)
        .str.replace(r"[^0-9.\-]", "", regex=True)
        .replace("", np.nan)
        .replace("nan", np.nan)
        .astype(float)
    )


def label(var: str) -> str:
    return EN_LABELS.get(var, var)


def safe_name(text: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", str(text)).strip("_")[:140]


def significance_stars(p: float) -> str:
    if pd.isna(p):
        return ""
    if p < 0.001:
        return "***"
    if p < 0.01:
        return "**"
    if p < 0.05:
        return "*"
    return "ns"


def p_text(p: float) -> str:
    if pd.isna(p):
        return "p = NA"
    if p < 0.001:
        return "p < 0.001"
    return f"p = {p:.3f}"


def strength_text(r: float) -> str:
    if pd.isna(r):
        return "undetermined"
    a = abs(r)
    if a < 0.10:
        return "negligible"
    if a < 0.30:
        return "weak"
    if a < 0.50:
        return "moderate"
    if a < 0.70:
        return "strong"
    return "very strong"


def direction_text(r: float) -> str:
    if pd.isna(r):
        return "undetermined"
    return "positive" if r >= 0 else "negative"


def shapiro_test(x: pd.Series) -> dict:
    x = pd.to_numeric(x, errors="coerce").dropna()
    if len(x) < 3:
        return {"n": len(x), "W": np.nan, "p": np.nan, "Distribution": "not testable"}
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        w, p = stats.shapiro(x)
    return {"n": len(x), "W": w, "p": p, "Distribution": "normal" if p > ALPHA else "non-normal"}


def sinclair_score(row: pd.Series, lift_col: str) -> float:
    sex = row.get("Pohlaví")
    bw = row.get("B.weight")
    lift = row.get(lift_col)
    if pd.isna(lift) or pd.isna(bw) or sex not in SINCLAIR_PARAMS:
        return np.nan
    A = SINCLAIR_PARAMS[sex]["A"]
    b = SINCLAIR_PARAMS[sex]["b"]
    coeff = 1.0 if bw > b else 10 ** (A * (np.log10(b / bw) ** 2))
    return lift * coeff


def corr_test(df: pd.DataFrame, x_col: str, y_col: str, group_name: str, hypothesis: str) -> dict:
    sub = df[[x_col, y_col]].apply(pd.to_numeric, errors="coerce").dropna()
    n = len(sub)
    base = {
        "Hypothesis": hypothesis,
        "Group": group_name,
        "Predictor": label(x_col),
        "Outcome": label(y_col),
        "Predictor_raw": x_col,
        "Outcome_raw": y_col,
        "n": n,
    }
    if n < 3 or sub[x_col].nunique() < 2 or sub[y_col].nunique() < 2:
        return {**base, "Normality predictor": "not testable", "Normality outcome": "not testable",
                "Method": "not applicable", "Coefficient": np.nan, "R2": np.nan, "p-value": np.nan,
                "Significance": "", "Result": "insufficient data"}

    norm_x = shapiro_test(sub[x_col])["Distribution"] == "normal"
    norm_y = shapiro_test(sub[y_col])["Distribution"] == "normal"
    if norm_x and norm_y:
        method = "Pearson r"
        coef, p = stats.pearsonr(sub[x_col], sub[y_col])
    else:
        method = "Spearman rho"
        coef, p = stats.spearmanr(sub[x_col], sub[y_col])

    return {**base,
            "Normality predictor": "normal" if norm_x else "non-normal",
            "Normality outcome": "normal" if norm_y else "non-normal",
            "Method": method,
            "Coefficient": coef,
            "R2": coef ** 2 if pd.notna(coef) else np.nan,
            "p-value": p,
            "Significance": significance_stars(p),
            "Result": "statistically significant" if pd.notna(p) and p < ALPHA else "not statistically significant"}


def group_compare(df: pd.DataFrame, value_col: str, group_name: str) -> dict:
    sub = df[[value_col, "Elite"]].copy()
    sub[value_col] = pd.to_numeric(sub[value_col], errors="coerce")
    sub = sub.dropna()
    elite = sub.loc[sub["Elite"], value_col]
    others = sub.loc[~sub["Elite"], value_col]
    base = {"Hypothesis": "H4", "Group": group_name, "Variable": label(value_col), "Variable_raw": value_col,
            "n Elite": len(elite), "n Others": len(others)}
    if len(elite) < 3 or len(others) < 3:
        return {**base, "Mean Elite": np.nan, "Mean Others": np.nan, "Median Elite": np.nan, "Median Others": np.nan,
                "Method": "not applicable", "Statistic": np.nan, "p-value": np.nan, "Significance": "", "Result": "insufficient data"}

    norm_e = shapiro_test(elite)["Distribution"] == "normal"
    norm_o = shapiro_test(others)["Distribution"] == "normal"
    if norm_e and norm_o:
        method = "Welch independent samples t-test"
        stat, p = stats.ttest_ind(elite, others, equal_var=False, nan_policy="omit")
    else:
        method = "Mann-Whitney U test"
        stat, p = stats.mannwhitneyu(elite, others, alternative="two-sided")

    return {**base,
            "Mean Elite": elite.mean(), "Mean Others": others.mean(),
            "Median Elite": elite.median(), "Median Others": others.median(),
            "SD Elite": elite.std(ddof=1), "SD Others": others.std(ddof=1),
            "Normality Elite": "normal" if norm_e else "non-normal",
            "Normality Others": "normal" if norm_o else "non-normal",
            "Method": method, "Statistic": stat, "p-value": p,
            "Significance": significance_stars(p),
            "Result": "statistically significant" if p < ALPHA else "not statistically significant"}


def add_regression_line(ax, x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    mask = np.isfinite(x) & np.isfinite(y)
    if mask.sum() < 3 or len(np.unique(x[mask])) < 2:
        return
    slope, intercept = np.polyfit(x[mask], y[mask], 1)
    xx = np.linspace(np.nanmin(x[mask]), np.nanmax(x[mask]), 100)
    ax.plot(xx, slope * xx + intercept, linewidth=2)


def plot_scatter(df: pd.DataFrame, x_col: str, y_col: str, title: str, out_path: Path, stats_row: dict | None = None):
    sub = df[[x_col, y_col, "Sex_label", "Age_group_label"]].copy()
    sub[x_col] = pd.to_numeric(sub[x_col], errors="coerce")
    sub[y_col] = pd.to_numeric(sub[y_col], errors="coerce")
    sub = sub.dropna()
    if len(sub) < 3:
        return False

    fig, ax = plt.subplots(figsize=(7.5, 5.5))
    for group_label, part in sub.groupby(["Sex_label", "Age_group_label"]):
        ax.scatter(part[x_col], part[y_col], alpha=0.75, label=" ".join(group_label), s=36)

    add_regression_line(ax, sub[x_col], sub[y_col])
    ax.set_xlabel(label(x_col))
    ax.set_ylabel(label(y_col))
    ax.set_title(title)
    if stats_row is not None and pd.notna(stats_row.get("Coefficient", np.nan)):
        annotation = (f"{stats_row['Method']}: {stats_row['Coefficient']:.2f}\n"
                      f"R² = {stats_row['R2']:.2f}, {p_text(stats_row['p-value'])} {stats_row['Significance']}")
        ax.text(0.03, 0.97, annotation, transform=ax.transAxes, va="top", ha="left",
                bbox=dict(boxstyle="round,pad=0.35", facecolor="white", alpha=0.85, edgecolor="0.75"))
    ax.legend(fontsize=8, frameon=True)
    fig.tight_layout()
    fig.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return True


def plot_boxplot(df: pd.DataFrame, value_col: str, title: str, out_path: Path, h4_row: dict | None = None):
    sub = df[[value_col, "Elite"]].copy()
    sub[value_col] = pd.to_numeric(sub[value_col], errors="coerce")
    sub = sub.dropna()
    if len(sub) < 3:
        return False
    data = [sub.loc[~sub["Elite"], value_col], sub.loc[sub["Elite"], value_col]]
    fig, ax = plt.subplots(figsize=(6.5, 5.5))
    
    # Matplotlib compatibility fix: Avoid 'labels' vs 'tick_labels' deprecation
    ax.boxplot(data, showmeans=True)
    ax.set_xticks([1, 2])
    ax.set_xticklabels(["Others", "Elite"])
    
    # add light jittered points
    for i, vals in enumerate(data, start=1):
        jitter = np.random.normal(i, 0.035, size=len(vals))
        ax.scatter(jitter, vals, alpha=0.55, s=24)
    ax.set_ylabel(label(value_col))
    ax.set_title(title)
    if h4_row is not None and pd.notna(h4_row.get("p-value", np.nan)):
        annotation = f"{h4_row['Method']}\n{p_text(h4_row['p-value'])} {h4_row['Significance']}"
        ax.text(0.03, 0.97, annotation, transform=ax.transAxes, va="top", ha="left",
                bbox=dict(boxstyle="round,pad=0.35", facecolor="white", alpha=0.85, edgecolor="0.75"))
    fig.tight_layout()
    fig.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return True


def interpretation_sentence(row: pd.Series) -> str:
    if row.get("Result") == "insufficient data":
        return f"In {row['Group']}, the relationship between {row['Predictor']} and {row['Outcome']} could not be tested due to insufficient data."
    coef_name = "r" if "Pearson" in row["Method"] else "rho"
    sig = "statistically significant" if row["p-value"] < ALPHA else "not statistically significant"
    return (f"In {row['Group']}, {row['Predictor']} showed a {strength_text(row['Coefficient'])} "
            f"{direction_text(row['Coefficient'])} association with {row['Outcome']} "
            f"({row['Method']}, {coef_name} = {row['Coefficient']:.2f}, R² = {row['R2']:.2f}, "
            f"{p_text(row['p-value'])}), which was {sig}.")


def comparison_sentence(row: pd.Series) -> str:
    if row.get("Result") == "insufficient data":
        return f"In {row['Group']}, {row['Variable']} could not be compared between elite and non-elite athletes due to insufficient data."
    sig = "statistically significant" if row["p-value"] < ALPHA else "not statistically significant"
    direction = "lower" if row["Mean Elite"] < row["Mean Others"] else "higher"
    return (f"In {row['Group']}, elite athletes had {direction} mean {row['Variable']} than the remaining athletes "
            f"({row['Mean Elite']:.2f} vs {row['Mean Others']:.2f}); this difference was {sig} "
            f"({row['Method']}, {p_text(row['p-value'])}).")


def style_excel(path: Path):
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    sig_fill = PatternFill("solid", fgColor="D9EAD3")
    ns_fill = PatternFill("solid", fgColor="F4CCCC")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # auto-width and wrap
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 45)
        # highlight Significance cells
        headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
        if "Significance" in headers:
            sig_col = headers["Significance"]
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=sig_col).value
                if val in ("*", "**", "***"):
                    ws.cell(row=row, column=sig_col).fill = sig_fill
                elif val == "ns":
                    ws.cell(row=row, column=sig_col).fill = ns_fill
    wb.save(path)

# =========================
# 3) Load and prepare data
# =========================
if not INPUT_FILE.exists():
    raise FileNotFoundError(f"Input file not found: {INPUT_FILE.resolve()}")

df = pd.read_csv(INPUT_FILE)
original_columns = df.columns.tolist()

numeric_cols = [
    "B.weight", "Snatch", "Clean Jerk", "Result", "Délka ruky (cm)", "Délka nohy (cm)",
    "Biceps (cm)", "Stehno (cm)", "Výška (cm)", "Handgrip P", "Handgrip L",
    "CMJ Height", "CMJ Force", "CMJ RFD", "CMJ Power", "SJ Height", "SJ Power", "SJ Force",
    "Weight", "Sinclair"
]
for col in numeric_cols:
    if col in df.columns:
        df[col] = to_numeric_series(df[col])

# English grouping labels
sex_map = {"Muži": "Men", "Ženy": "Women"}
age_map = {"Junior": "U20", "U20": "U20", "U23": "U23"}
df["Sex_label"] = df["Pohlaví"].map(sex_map).fillna(df["Pohlaví"].astype(str))
df["Age_group_label"] = df["Věková skupina"].map(age_map).fillna(df["Věková skupina"].astype(str))
df["Research group"] = df["Sex_label"] + " " + df["Age_group_label"]

# Calculated indices
df["EUR"] = df["CMJ Height"] / df["SJ Height"]
df["RLL"] = df["Délka nohy (cm)"] / df["Výška (cm)"] * 100
df["RAL"] = df["Délka ruky (cm)"] / df["Výška (cm)"] * 100

# Dominant hand was not specified in the dataset; maximum of left/right is used as a proxy.
df["Dominant handgrip"] = df[["Handgrip P", "Handgrip L"]].max(axis=1)

# Sinclair scores for individual lifts and total
df["Snatch Sinclair"] = df.apply(lambda r: sinclair_score(r, "Snatch"), axis=1)
df["Clean Jerk Sinclair"] = df.apply(lambda r: sinclair_score(r, "Clean Jerk"), axis=1)
df["Total Sinclair calculated"] = df.apply(lambda r: sinclair_score(r, "Result"), axis=1)

# Elite group: top 25% within each sex based on the source Sinclair score
cutoffs = df.groupby("Pohlaví")["Sinclair"].quantile(0.75).to_dict()
df["Elite cutoff"] = df["Pohlaví"].map(cutoffs)
df["Elite"] = df["Sinclair"] >= df["Elite cutoff"]

performance_abs = ["Snatch", "Clean Jerk", "Result"]
performance_rel = ["Snatch Sinclair", "Clean Jerk Sinclair", "Total Sinclair calculated", "Sinclair"]
performance_all = performance_abs + performance_rel

groups = {name: g.copy() for name, g in df.groupby("Research group", dropna=False)}
groups["Whole sample"] = df.copy()

# =========================
# 4) Normality testing
# =========================
normality_vars = sorted(set([
    "CMJ Height", "SJ Height", "Dominant handgrip", "Biceps (cm)", "Stehno (cm)",
    "EUR", "RLL", "RAL"
] + performance_all))

normality_rows = []
for group_name, g in groups.items():
    for col in normality_vars:
        if col in g.columns:
            res = shapiro_test(g[col])
            normality_rows.append({"Group": group_name, "Variable": label(col), "Variable_raw": col, **res})
normality_df = pd.DataFrame(normality_rows)
normality_df["Significance"] = normality_df["p"].apply(significance_stars)

# =========================
# 5) Hypothesis tests
# =========================
correlation_rows = []
for group_name, g in groups.items():
    # H1: CMJ/SJ height vs absolute and relative competitive performance
    for x in ["CMJ Height", "SJ Height"]:
        for y in performance_all:
            correlation_rows.append(corr_test(g, x, y, group_name, "H1"))

    # H2: dominant handgrip vs competitive success
    for y in performance_all:
        correlation_rows.append(corr_test(g, "Dominant handgrip", y, group_name, "H2"))

    # H3: arm and thigh circumferences vs performance
    for x in ["Biceps (cm)", "Stehno (cm)"]:
        for y in performance_all:
            correlation_rows.append(corr_test(g, x, y, group_name, "H3"))

    # H5: EUR vs Sinclair success
    for y in ["Sinclair", "Total Sinclair calculated"]:
        correlation_rows.append(corr_test(g, "EUR", y, group_name, "H5"))

corr_df = pd.DataFrame(correlation_rows)
corr_df["Interpretation sentence"] = corr_df.apply(interpretation_sentence, axis=1)

# H2 Snatch vs Clean & Jerk comparison
h2_compare = (
    corr_df[(corr_df["Hypothesis"] == "H2") &
            (corr_df["Outcome_raw"].isin(["Snatch", "Clean Jerk", "Snatch Sinclair", "Clean Jerk Sinclair"]))]
    .pivot_table(index=["Group"], columns="Outcome_raw", values="Coefficient", aggfunc="first")
    .reset_index()
)
if {"Snatch", "Clean Jerk"}.issubset(h2_compare.columns):
    h2_compare["Absolute performance: |Snatch| > |Clean & Jerk|"] = h2_compare["Snatch"].abs() > h2_compare["Clean Jerk"].abs()
if {"Snatch Sinclair", "Clean Jerk Sinclair"}.issubset(h2_compare.columns):
    h2_compare["Relative performance: |Snatch Sinclair| > |Clean & Jerk Sinclair|"] = h2_compare["Snatch Sinclair"].abs() > h2_compare["Clean Jerk Sinclair"].abs()

# H3 absolute vs relative strength comparison
h3_summary_rows = []
for (group_name, x_raw), sub in corr_df[corr_df["Hypothesis"] == "H3"].groupby(["Group", "Predictor_raw"]):
    abs_mean = sub[sub["Outcome_raw"].isin(performance_abs)]["Coefficient"].abs().mean()
    rel_mean = sub[sub["Outcome_raw"].isin(performance_rel)]["Coefficient"].abs().mean()
    h3_summary_rows.append({
        "Group": group_name,
        "Variable": label(x_raw),
        "Mean absolute coefficient for kg performance": abs_mean,
        "Mean absolute coefficient for Sinclair performance": rel_mean,
        "Stronger relationship with absolute performance": bool(abs_mean > rel_mean) if pd.notna(abs_mean) and pd.notna(rel_mean) else np.nan,
    })
h3_summary = pd.DataFrame(h3_summary_rows)

# H4 group comparison
h4_rows = []
for group_name, g in groups.items():
    for var in ["RLL", "RAL"]:
        h4_rows.append(group_compare(g, var, group_name))
h4_df = pd.DataFrame(h4_rows)
h4_df["Interpretation sentence"] = h4_df.apply(comparison_sentence, axis=1)

# =========================
# 6) Publication tables
# =========================
def rounded(df_in: pd.DataFrame, ndigits: int = 3) -> pd.DataFrame:
    out = df_in.copy()
    for col in out.select_dtypes(include=[np.number]).columns:
        out[col] = out[col].round(ndigits)
    return out

h1_table = rounded(corr_df[corr_df["Hypothesis"] == "H1"].drop(columns=["Predictor_raw", "Outcome_raw"]))
h2_table = rounded(corr_df[corr_df["Hypothesis"] == "H2"].drop(columns=["Predictor_raw", "Outcome_raw"]))
h3_table = rounded(corr_df[corr_df["Hypothesis"] == "H3"].drop(columns=["Predictor_raw", "Outcome_raw"]))
h4_table = rounded(h4_df.drop(columns=["Variable_raw"]))
h5_table = rounded(corr_df[corr_df["Hypothesis"] == "H5"].drop(columns=["Predictor_raw", "Outcome_raw"]))

for name, table in {
    "Table_H1_vertical_jump_correlations.csv": h1_table,
    "Table_H2_handgrip_correlations.csv": h2_table,
    "Table_H2_snatch_vs_cleanjerk_comparison.csv": rounded(h2_compare),
    "Table_H3_circumference_correlations.csv": h3_table,
    "Table_H3_absolute_vs_relative_summary.csv": rounded(h3_summary),
    "Table_H4_elite_vs_others.csv": h4_table,
    "Table_H5_EUR_correlations.csv": h5_table,
}.items():
    table.to_csv(TABLE_DIR / name, index=False, encoding="utf-8-sig")

# =========================
# 7) Figures (Now Generated Separately Per Group)
# =========================
def get_stat(group_name: str, hypothesis: str, x_raw: str, y_raw: str):
    m = corr_df[(corr_df["Group"] == group_name) &
                (corr_df["Hypothesis"] == hypothesis) &
                (corr_df["Predictor_raw"] == x_raw) &
                (corr_df["Outcome_raw"] == y_raw)]
    return None if m.empty else m.iloc[0].to_dict()

def get_h4_stat(group_name: str, var: str):
    m = h4_df[(h4_df["Group"] == group_name) & (h4_df["Variable_raw"] == var)]
    return None if m.empty else m.iloc[0].to_dict()

# Generate graphs iteratively for EVERY group (U20 Men, U23 Men, U20 Women, U23 Women, plus Whole Sample)
for group_name, group_df in groups.items():
    safe_group = safe_name(group_name)

    # H1: key plots
    for x in ["CMJ Height", "SJ Height"]:
        for y in ["Result", "Sinclair"]:
            plot_scatter(
                group_df, x, y,
                f"H1 ({group_name}): {label(x)} and {label(y)}",
                FIGURE_DIR / "H1" / f"H1_{safe_name(x)}_vs_{safe_name(y)}_{safe_group}.png",
                get_stat(group_name, "H1", x, y)
            )

    # H2: key plots + Snatch/Clean & Jerk
    for y in ["Snatch", "Clean Jerk", "Sinclair"]:
        plot_scatter(
            group_df, "Dominant handgrip", y,
            f"H2 ({group_name}): {label('Dominant handgrip')} and {label(y)}",
            FIGURE_DIR / "H2" / f"H2_Dominant_handgrip_vs_{safe_name(y)}_{safe_group}.png",
            get_stat(group_name, "H2", "Dominant handgrip", y)
        )

    # H3: key plots
    for x in ["Biceps (cm)", "Stehno (cm)"]:
        for y in ["Result", "Sinclair"]:
            plot_scatter(
                group_df, x, y,
                f"H3 ({group_name}): {label(x)} and {label(y)}",
                FIGURE_DIR / "H3" / f"H3_{safe_name(x)}_vs_{safe_name(y)}_{safe_group}.png",
                get_stat(group_name, "H3", x, y)
            )

    # H4: boxplots for RLL and RAL
    for var in ["RLL", "RAL"]:
        plot_boxplot(
            group_df, var,
            f"H4 ({group_name}): {label(var)} in elite and non-elite athletes",
            FIGURE_DIR / "H4" / f"H4_{safe_name(var)}_elite_vs_others_{safe_group}.png",
            get_h4_stat(group_name, var)
        )

    # H5: EUR vs Sinclair
    for y in ["Sinclair", "Total Sinclair calculated"]:
        plot_scatter(
            group_df, "EUR", y,
            f"H5 ({group_name}): {label('EUR')} and {label(y)}",
            FIGURE_DIR / "H5" / f"H5_EUR_vs_{safe_name(y)}_{safe_group}.png",
            get_stat(group_name, "H5", "EUR", y)
        )

# =========================
# 8) Thesis interpretation text
# =========================
lines = []
lines.append("STATISTICAL SIGNIFICANCE CODING")
lines.append("ns = p >= 0.05; * = p < 0.05; ** = p < 0.01; *** = p < 0.001")
lines.append("")
lines.append("NOTE ON METHODS")
lines.append("Normality was assessed using the Shapiro-Wilk test. Pearson's correlation coefficient was used when both variables were normally distributed; otherwise, Spearman's rho was used. For H4, Welch's independent samples t-test was used when both groups were normally distributed; otherwise, the Mann-Whitney U test was used.")
lines.append("")
lines.append("NOTE ON HANDGRIP")
lines.append("Dominant handgrip strength was approximated as the higher value from the right and left handgrip measurements because hand dominance was not explicitly available in the dataset.")
lines.append("")
for h in ["H1", "H2", "H3", "H5"]:
    lines.append(f"{h} - selected whole-sample interpretation sentences")
    sub = corr_df[(corr_df["Hypothesis"] == h) & (corr_df["Group"] == "Whole sample")].copy()
    # Put statistically significant and stronger effects first, but keep all tested relationships.
    sub["abs_coef"] = sub["Coefficient"].abs()
    sub = sub.sort_values(["p-value", "abs_coef"], ascending=[True, False])
    for sentence in sub["Interpretation sentence"].head(12):
        lines.append("- " + sentence)
    lines.append("")
lines.append("H4 - whole-sample interpretation sentences")
for sentence in h4_df[h4_df["Group"] == "Whole sample"]["Interpretation sentence"]:
    lines.append("- " + sentence)
lines.append("")
lines.append("Suggested reporting strategy")
lines.append("Use one comprehensive correlation table for each hypothesis and include only the most informative figures in the thesis text. For H4, boxplots are recommended because the hypothesis compares two groups. For H1, H2, H3 and H5, scatter plots with regression lines are recommended for the strongest or theoretically most relevant associations.")
INTERPRETATION_TXT.write_text("\n".join(lines), encoding="utf-8")

# =========================
# 9) Excel export and styling
# =========================
with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
    rounded(df).to_excel(writer, sheet_name="Calculated_data", index=False)
    rounded(normality_df).to_excel(writer, sheet_name="Table_1_Normality", index=False)
    h1_table.to_excel(writer, sheet_name="Table_2_H1", index=False)
    h2_table.to_excel(writer, sheet_name="Table_3_H2", index=False)
    rounded(h2_compare).to_excel(writer, sheet_name="H2_Snatch_vs_CJ", index=False)
    h3_table.to_excel(writer, sheet_name="Table_4_H3", index=False)
    rounded(h3_summary).to_excel(writer, sheet_name="H3_abs_vs_relative", index=False)
    h4_table.to_excel(writer, sheet_name="Table_5_H4", index=False)
    h5_table.to_excel(writer, sheet_name="Table_6_H5", index=False)

style_excel(OUTPUT_XLSX)

# =========================
# 10) Log
# =========================
log = f"""
Analysis completed successfully.
Input file: {INPUT_FILE.resolve()}
Rows: {len(df)}
Original columns: {original_columns}

Main outputs:
- Excel workbook: {OUTPUT_XLSX.resolve()}
- Publication CSV tables: {TABLE_DIR.resolve()}
- Figures: {FIGURE_DIR.resolve()}
- Interpretation text: {INTERPRETATION_TXT.resolve()}

Important notes:
1) Dominant handgrip = max(Handgrip P, Handgrip L), because explicit hand dominance is not available.
2) Elite group = top 25% within each sex based on the source Sinclair score.
3) Significance coding: ns = p >= 0.05; * = p < 0.05; ** = p < 0.01; *** = p < 0.001.
""".strip()
LOG_TXT.write_text(log, encoding="utf-8")

print(log)